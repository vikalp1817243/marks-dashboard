from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import func

from database import get_db
from models import ExamSession, Submission, ExamScore
from schemas import SessionCreate, SessionResponse
from auth import verify_google_token
from datetime import datetime
from uuid import uuid4
import csv
import io
import openpyxl
from odf.opendocument import load as load_ods
from odf.table import Table, TableRow, TableCell
from odf.text import P
from odf.teletype import extractText

from config import BASE_URL
from stats import recalculate_stats
from websocket_manager import manager

router = APIRouter(prefix="/api/sessions", tags=["Sessions"])

# ---------------------------------------------------------------------------
# Column-header keywords the parser searches for (case-insensitive).
# The FIRST matching column in the header row wins.
# ---------------------------------------------------------------------------
MARKS_HEADER_KEYWORDS = {
    'marks', 'mark', 'score', 'scores', 'total', 'obtained',
    'total marks', 'obtained marks', 'total_marks', 'obtained_marks',
}


def _parse_ods_row(table_row) -> list:
    """Parse an ODS TableRow, correctly expanding cells that have
    the table:number-columns-repeated attribute."""
    cells = []
    for cell in table_row.getElementsByType(TableCell):
        text = extractText(cell).strip()
        value = text if text else None
        # Handle repeated cells (e.g. empty columns)
        repeat = cell.getAttribute('numbercolumnsrepeated')
        if repeat:
            try:
                repeat_count = int(repeat)
            except (ValueError, TypeError):
                repeat_count = 1
            # Cap repeat expansion to avoid memory issues from huge padding
            repeat_count = min(repeat_count, 200)
            cells.extend([value] * repeat_count)
        else:
            cells.append(value)
    return cells


def _find_marks_column(header_cells: list) -> int | None:
    """Return the 0-based index of the first column whose header matches
    one of the MARKS_HEADER_KEYWORDS (case-insensitive substring match)."""
    for idx, cell in enumerate(header_cells):
        if cell is None:
            continue
        cleaned = str(cell).strip().lower()
        if cleaned in MARKS_HEADER_KEYWORDS:
            return idx
    return None


def _extract_scores_from_rows(rows: list[list], max_marks: float) -> list[float]:
    """Given a list of rows (each row = list of cell values), find the marks
    column via header detection and return a list of valid scores.
    Falls back to scanning each row for the first numeric value."""
    if not rows:
        return []

    # --- Strategy 1: Header-based column detection ---
    header = rows[0]
    marks_col = _find_marks_column(header)

    scores = []
    if marks_col is not None:
        # Extract strictly from the identified column
        for row in rows[1:]:
            if marks_col >= len(row):
                continue
            cell_val = row[marks_col]
            if cell_val is None:
                continue
            try:
                val = float(str(cell_val).strip())
            except (ValueError, TypeError):
                continue
            if 0 <= val <= max_marks:
                scores.append(val)
        if scores:
            return scores

    # --- Strategy 2: Fallback – first numeric value per row ---
    for row in rows:
        for cell in row:
            if cell is None:
                continue
            try:
                val = float(str(cell).strip())
            except (ValueError, TypeError):
                continue
            if 0 <= val <= max_marks:
                scores.append(val)
                break  # one score per row

    return scores


@router.get("")
async def get_all_sessions(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(ExamSession).order_by(ExamSession.created_at.desc()).limit(100)
    )
    sessions = result.scalars().all()

    active = []
    expired = []
    now = datetime.utcnow()

    for session_obj in sessions:
        count_res = await db.execute(
            select(func.count(Submission.id)).where(
                Submission.session_id == session_obj.id
            )
        )
        sub_count = count_res.scalar()

        s_dict = {
            "id": session_obj.id,
            "name": session_obj.name,
            "max_marks": session_obj.max_marks,
            "class_size": session_obj.class_size,
            "submissions": sub_count,
            "spots_remaining": session_obj.class_size - sub_count,
            "created_at": session_obj.created_at,
            "expires_at": session_obj.expires_at,
            "submission_url": f"{BASE_URL}/submit.html?session={session_obj.id}",
            "dashboard_url": f"{BASE_URL}/dashboard.html?session={session_obj.id}",
        }
        if session_obj.expires_at > now:
            active.append(s_dict)
        else:
            expired.append(s_dict)

    return {"active": active, "expired": expired}


@router.post("", response_model=SessionResponse, status_code=status.HTTP_201_CREATED)
async def create_session(
    session_data: SessionCreate,
    email: str = Depends(verify_google_token),
    db: AsyncSession = Depends(get_db),
):
    # class_size is required for the manual entry path
    if session_data.class_size is None:
        raise HTTPException(
            status_code=400, detail="class_size is required for manual session creation"
        )
    if not (10 <= session_data.class_size <= 1000):
        raise HTTPException(
            status_code=400, detail="class_size must be between 10 and 1000"
        )

    max_marks = 50 if "50" in session_data.exam_type else 100

    c_id = session_data.class_id.strip() if session_data.class_id else ""
    fac = session_data.faculty_name.strip()
    slot = session_data.slot.strip()
    course = session_data.course_code.strip()

    if c_id:
        unique_identifier = f"{c_id}_{fac}_{slot}".upper()
    else:
        unique_identifier = f"{fac}_{slot}_{course}".upper()

    session_name = f"{session_data.exam_type} - {unique_identifier}"

    existing_res = await db.execute(
        select(ExamSession).where(ExamSession.unique_identifier == unique_identifier)
    )
    existing_session = existing_res.scalar_one_or_none()
    base_url = BASE_URL

    if existing_session:
        count_res = await db.execute(
            select(func.count(Submission.id)).where(
                Submission.session_id == existing_session.id
            )
        )
        count = count_res.scalar()

        return SessionResponse(
            id=existing_session.id,
            name=existing_session.name,
            max_marks=existing_session.max_marks,
            class_size=existing_session.class_size,
            created_at=existing_session.created_at,
            expires_at=existing_session.expires_at,
            spots_remaining=existing_session.class_size - count,
            submission_url=f"{base_url}/submit.html?session={existing_session.id}",
            dashboard_url=f"{base_url}/dashboard.html?session={existing_session.id}",
        )

    new_session = ExamSession(
        name=session_name,
        max_marks=max_marks,
        class_size=session_data.class_size,
        unique_identifier=unique_identifier,
    )
    db.add(new_session)
    await db.commit()
    await db.refresh(new_session)

    return SessionResponse(
        id=new_session.id,
        name=new_session.name,
        max_marks=new_session.max_marks,
        class_size=new_session.class_size,
        created_at=new_session.created_at,
        expires_at=new_session.expires_at,
        spots_remaining=new_session.class_size,
        submission_url=f"{base_url}/submit.html?session={new_session.id}",
        dashboard_url=f"{base_url}/dashboard.html?session={new_session.id}",
    )


@router.post("/bulk-create")
async def bulk_create_session(
    exam_type: str = Form(...),
    class_id: str = Form(""),
    faculty_name: str = Form(...),
    slot: str = Form(...),
    course_code: str = Form(...),
    file: UploadFile = File(...),
    email: str = Depends(verify_google_token),
    db: AsyncSession = Depends(get_db),
):
    """Atomic endpoint: parse file FIRST, reject if no valid column header,
    then create session with class_size = len(scores) and insert all scores.
    No database records are created if the file is invalid."""

    max_marks = 50 if "50" in exam_type else 100

    # ── 1. Parse file BEFORE touching the database ──────────────────────
    contents = await file.read()
    rows: list[list] = []

    if file.filename.endswith(".csv"):
        text_content = contents.decode("utf-8-sig")
        rows = [row for row in csv.reader(text_content.splitlines()) if row]
    elif file.filename.endswith((".xlsx", ".xls")):
        wb = openpyxl.load_workbook(
            filename=io.BytesIO(contents), data_only=True
        )
        sheet = wb.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True) if any(c is not None for c in row)]
    elif file.filename.endswith(".ods"):
        doc = load_ods(io.BytesIO(contents))
        for sheet in doc.spreadsheet.getElementsByType(Table):
            for table_row in sheet.getElementsByType(TableRow):
                cells = _parse_ods_row(table_row)
                if any(c is not None for c in cells):
                    rows.append(cells)
    else:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file format. Please upload .csv, .xlsx, or .ods",
        )

    # Require a recognized column header — fallback not allowed for bulk-create
    if not rows:
        raise HTTPException(status_code=400, detail="The file appears to be empty.")

    header = rows[0]
    marks_col = _find_marks_column(header)

    if marks_col is None:
        raise HTTPException(
            status_code=400,
            detail="No column header named 'Marks', 'Score', 'Total', or 'Obtained' found "
            "in the first row. Please add a header and re-upload.",
        )

    # Extract ALL numeric values from the identified column first
    all_numeric = []
    for row in rows[1:]:
        if marks_col >= len(row):
            continue
        cell_val = row[marks_col]
        if cell_val is None:
            continue
        try:
            val = float(str(cell_val).strip())
        except (ValueError, TypeError):
            continue
        if val >= 0:
            all_numeric.append(val)

    if not all_numeric:
        raise HTTPException(
            status_code=400,
            detail=f"Found the '{header[marks_col]}' column but no valid numeric scores were found.",
        )

    # Pre-validation: for Mid-Term (50 marks), reject file if ANY score exceeds 50
    over_limit = [v for v in all_numeric if v > max_marks]
    if over_limit:
        raise HTTPException(
            status_code=400,
            detail=f"Your file contains {len(over_limit)} score(s) exceeding the "
            f"{max_marks}-mark limit (e.g. {over_limit[0]:.1f}). "
            f"For a {'Mid-Term (50 marks)' if max_marks == 50 else 'Term-End (100 marks)'} session, "
            f"all marks must be between 0 and {max_marks}. "
            f"Please correct the file and re-upload.",
        )

    scores = all_numeric

    # ── 2. File is valid — now create session ───────────────────────────
    c_id = class_id.strip() if class_id else ""
    fac = faculty_name.strip()
    s = slot.strip()
    course = course_code.strip()

    if c_id:
        unique_identifier = f"{c_id}_{fac}_{s}".upper()
    else:
        unique_identifier = f"{fac}_{s}_{course}".upper()

    session_name = f"{exam_type} - {unique_identifier}"
    class_size = len(scores)
    base_url = BASE_URL

    # Check for existing session
    existing_res = await db.execute(
        select(ExamSession).where(ExamSession.unique_identifier == unique_identifier)
    )
    existing_session = existing_res.scalar_one_or_none()

    if existing_session:
        # Update class_size to match new file and clear old scores
        existing_session.class_size = class_size
        await db.commit()
        session_obj = existing_session
    else:
        session_obj = ExamSession(
            name=session_name,
            max_marks=max_marks,
            class_size=class_size,
            unique_identifier=unique_identifier,
        )
        db.add(session_obj)
        await db.commit()
        await db.refresh(session_obj)

    # ── 3. Insert all scores ────────────────────────────────────────────
    for val in scores:
        db.add(ExamScore(session_id=session_obj.id, score_value=round(val, 3)))
        db.add(
            Submission(
                session_id=session_obj.id, hashed_student_id=f"bulk_{uuid4().hex}"
            )
        )
    await db.commit()

    # ── 4. Trigger stats recalculation ──────────────────────────────────
    new_stats_model = await recalculate_stats(session_obj.id, db)
    if new_stats_model:
        await manager.broadcast_stats(session_obj.id, new_stats_model)

    return {
        "id": session_obj.id,
        "name": session_obj.name,
        "max_marks": session_obj.max_marks,
        "class_size": class_size,
        "scores_imported": len(scores),
        "dashboard_url": f"{base_url}/dashboard.html?session={session_obj.id}",
    }

@router.post("/{session_id}/bulk")
async def bulk_upload_scores(
    session_id: str,
    file: UploadFile = File(...),
    email: str = Depends(verify_google_token),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(ExamSession).where(ExamSession.id == session_id)
    )
    session_obj = result.scalar_one_or_none()

    if not session_obj:
        raise HTTPException(status_code=404, detail="Session not found")

    if session_obj.expires_at < datetime.utcnow():
        raise HTTPException(status_code=410, detail="Session has expired")

    contents = await file.read()
    max_marks = session_obj.max_marks

    # ── Convert file → list[list[cell_value]] ──────────────────────────────
    rows: list[list] = []

    if file.filename.endswith(".csv"):
        text_content = contents.decode("utf-8-sig")
        rows = [row for row in csv.reader(text_content.splitlines()) if row]

    elif file.filename.endswith((".xlsx", ".xls")):
        wb = openpyxl.load_workbook(
            filename=io.BytesIO(contents), data_only=True
        )
        sheet = wb.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True) if any(c is not None for c in row)]

    elif file.filename.endswith(".ods"):
        doc = load_ods(io.BytesIO(contents))
        for sheet in doc.spreadsheet.getElementsByType(Table):
            for table_row in sheet.getElementsByType(TableRow):
                cells = _parse_ods_row(table_row)
                if any(c is not None for c in cells):
                    rows.append(cells)
    else:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file format. Please upload .csv, .xlsx, or .ods",
        )

    # ── Smart extraction via header detection + fallback ───────────────────
    scores = _extract_scores_from_rows(rows, max_marks)

    if not scores:
        raise HTTPException(
            status_code=400,
            detail="No readable numeric scores found in the spreadsheet. "
            "Tip: make sure there is a column header named 'Marks' or 'Score'.",
        )

    # ── Capacity check & insertion ─────────────────────────────────────────
    sub_count_res = await db.execute(
        select(func.count(Submission.id)).where(Submission.session_id == session_id)
    )
    current_count = sub_count_res.scalar()
    capacity = session_obj.class_size - current_count

    if capacity <= 0:
        raise HTTPException(status_code=400, detail="Session is already full")

    scores = scores[:capacity]

    for val in scores:
        db.add(ExamScore(session_id=session_id, score_value=round(val, 3)))
        db.add(
            Submission(
                session_id=session_id, hashed_student_id=f"bulk_{uuid4().hex}"
            )
        )

    await db.commit()

    new_stats_model = await recalculate_stats(session_id, db)
    if new_stats_model:
        await manager.broadcast_stats(session_id, new_stats_model)

    return {"message": f"Successfully imported {len(scores)} scores."}


@router.get("/{session_id}", response_model=SessionResponse)
async def get_session(session_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(ExamSession).where(ExamSession.id == session_id)
    )
    session_obj = result.scalar_one_or_none()

    if not session_obj:
        raise HTTPException(status_code=404, detail="Session not found")

    if session_obj.expires_at < datetime.utcnow():
        raise HTTPException(status_code=410, detail="Session has expired")

    count_result = await db.execute(
        select(func.count(Submission.id)).where(Submission.session_id == session_id)
    )
    current_count = count_result.scalar()

    base_url = BASE_URL
    return SessionResponse(
        id=session_obj.id,
        name=session_obj.name,
        max_marks=session_obj.max_marks,
        class_size=session_obj.class_size,
        created_at=session_obj.created_at,
        expires_at=session_obj.expires_at,
        spots_remaining=session_obj.class_size - current_count,
        submission_url=f"{base_url}/submit.html?session={session_obj.id}",
        dashboard_url=f"{base_url}/dashboard.html?session={session_obj.id}",
    )

